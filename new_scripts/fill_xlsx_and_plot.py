#!/usr/bin/env python3
"""Fill rebuttal_experiment_data.xlsx E4/E5-E6/Status sheets and generate plots."""
import json, sys
from pathlib import Path
from collections import defaultdict
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font

ROOT = Path("/common/users/sl2148/Public/yang_ouyang/projects/fact-enhancement")
MULTI_DIR = ROOT / "new_exps" / "figure1_multi_dataset"
GSM8K_DIR = ROOT / "new_exps" / "figure1_sampling_data"
XLSX_PATH = ROOT / "documents" / "rebuttal_experiment_data.xlsx"
PLOT_DIR  = ROOT / "documents" / "e4_plots"

EXPECTED = {"gsm8k": 1319, "math500": 500, "aime": 30, "amc": 40, "olympiad": 675}
DS_DISPLAY = {"gsm8k": "GSM8K", "math500": "MATH-500", "aime": "AIME", "amc": "AMC", "olympiad": "Olympiad"}

def get_jsonl_path(model_san, dataset):
    if dataset == "gsm8k":
        return GSM8K_DIR / model_san / "gsm8k_samples.jsonl"
    return MULTI_DIR / dataset / model_san / "samples.jsonl"


def load_samples(model_san, dataset):
    p = get_jsonl_path(model_san, dataset)
    if not p.is_file():
        return None
    records = []
    with open(p, encoding="utf-8") as f:
        for line in f:
            records.append(json.loads(line))
    return records


def compute_e4_stats(records):
    n_q = len(records)
    all_correct, all_steps, all_rho, all_total_tok = [], [], [], []
    for rec in records:
        for s in rec["samples"]:
            all_correct.append(1.0 if s["correct"] else 0.0)
            all_steps.append(s["n_steps"])
            all_rho.append(s["density_rho"])
            all_total_tok.append(s["total_tokens"])
    n_s = len(all_correct)
    acc = round(np.mean(all_correct) * 100, 2)
    std_acc = round(np.std([np.mean([1.0 if s["correct"] else 0.0 for s in rec["samples"]]) for rec in records]) * 100, 2)
    avg_steps = round(np.mean(all_steps), 2)
    std_steps = round(np.std(all_steps), 2)
    avg_rho = round(np.mean(all_rho), 2)
    std_rho = round(np.std(all_rho), 2)
    avg_tok = round(np.mean(all_total_tok), 2)
    return n_q, n_s, acc, std_acc, avg_steps, std_steps, avg_rho, std_rho, avg_tok


def compute_e56_stats(records):
    correct_steps, incorrect_steps = [], []
    correct_rho, incorrect_rho = [], []
    for rec in records:
        for s in rec["samples"]:
            if s["correct"]:
                correct_steps.append(s["n_steps"])
                correct_rho.append(s["density_rho"])
            else:
                incorrect_steps.append(s["n_steps"])
                incorrect_rho.append(s["density_rho"])
    n_c = len(correct_rho)
    n_i = len(incorrect_rho)
    if n_c == 0 or n_i == 0:
        return None
    c_avg_s = round(np.mean(correct_steps), 2)
    i_avg_s = round(np.mean(incorrect_steps), 2)
    d_s = round(c_avg_s - i_avg_s, 2)
    c_avg_r = round(np.mean(correct_rho), 2)
    i_avg_r = round(np.mean(incorrect_rho), 2)
    d_r = round(c_avg_r - i_avg_r, 2)
    overall_rho = round(np.mean(correct_rho + incorrect_rho), 2)
    overall_acc = round(n_c / (n_c + n_i) * 100, 2)
    return n_c, n_i, c_avg_s, i_avg_s, d_s, c_avg_r, i_avg_r, d_r, overall_rho, overall_acc


# ── E4 / E5-E6 row mapping ──────────────────────────────────────────────────
# (model_sanitized, model_display, params_b, datasets_in_order)
QWEN_MODELS = [
    ("Qwen_Qwen2.5-0.5B-Instruct", "Qwen2.5-0.5B-Instruct", 0.5),
    ("Qwen_Qwen2.5-1.5B-Instruct", "Qwen2.5-1.5B-Instruct", 1.5),
    ("Qwen_Qwen2.5-3B-Instruct",   "Qwen2.5-3B-Instruct",   3),
    ("Qwen_Qwen2.5-7B-Instruct",   "Qwen2.5-7B-Instruct",   7),
    ("Qwen_Qwen2.5-14B-Instruct",  "Qwen2.5-14B-Instruct",  14),
    ("Qwen_Qwen2.5-32B-Instruct",  "Qwen2.5-32B-Instruct",  32),
    ("Qwen_Qwen2.5-72B-Instruct",  "Qwen2.5-72B-Instruct",  72),
]
LLAMA_MODELS = [
    ("meta-llama_Llama-3.2-1B-Instruct", "Llama-3.2-1B-Instruct", 1),
    ("meta-llama_Llama-3.2-3B-Instruct", "Llama-3.2-3B-Instruct", 3),
    ("meta-llama_Llama-3.1-8B-Instruct", "Llama-3.1-8B-Instruct", 8),
    ("meta-llama_Llama-3.1-70B-Instruct","Llama-3.1-70B-Instruct",70),
]
ALL_MODELS = QWEN_MODELS + LLAMA_MODELS
DATASETS = ["gsm8k", "math500", "aime", "amc", "olympiad"]

STATUS_SHORT = {
    "Qwen_Qwen2.5-0.5B-Instruct": "Qwen2.5-0.5B",
    "Qwen_Qwen2.5-1.5B-Instruct": "Qwen2.5-1.5B",
    "Qwen_Qwen2.5-3B-Instruct":   "Qwen2.5-3B",
    "Qwen_Qwen2.5-7B-Instruct":   "Qwen2.5-7B",
    "Qwen_Qwen2.5-14B-Instruct":  "Qwen2.5-14B",
    "Qwen_Qwen2.5-32B-Instruct":  "Qwen2.5-32B",
    "Qwen_Qwen2.5-72B-Instruct":  "Qwen2.5-72B",
    "meta-llama_Llama-3.2-1B-Instruct": "Llama-3.2-1B",
    "meta-llama_Llama-3.2-3B-Instruct": "Llama-3.2-3B",
    "meta-llama_Llama-3.1-8B-Instruct": "Llama-3.1-8B",
    "meta-llama_Llama-3.1-70B-Instruct":"Llama-3.1-70B",
}


def fill_e4_sheet(wb):
    """Rebuild E4 Fig1 Scaling sheet from data files."""
    if "E4 Fig1 Scaling" in wb.sheetnames:
        idx = wb.sheetnames.index("E4 Fig1 Scaling")
        wb.remove(wb["E4 Fig1 Scaling"])
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet("E4 Fig1 Scaling", idx)

    headers = ["Model", "Params\n(B)", "Dataset", "N\nQuestions", "N\nSamples",
               "Accuracy\n(%)", "Std Dev\n(%)", "Avg\nSteps", "Std\nSteps",
               "Avg \u03c1\n(tok/step)", "Std \u03c1", "Avg Total\nTokens", "Status"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    filled = 0
    for model_list, family_label in [(QWEN_MODELS, None), (LLAMA_MODELS, "Llama Family")]:
        if family_label:
            ws.append([family_label])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        for model_san, model_disp, params_b in model_list:
            for ds in DATASETS:
                records = load_samples(model_san, ds)
                if records and len(records) >= EXPECTED[ds]:
                    n_q, n_s, acc, std_acc, avg_s, std_s, avg_r, std_r, avg_t = compute_e4_stats(records)
                    ws.append([model_disp, params_b, DS_DISPLAY[ds], n_q, n_s,
                               acc, std_acc, avg_s, std_s, avg_r, std_r, avg_t, "DONE"])
                    filled += 1
                else:
                    ws.append([model_disp, params_b, DS_DISPLAY[ds], EXPECTED[ds],
                               EXPECTED[ds]*16 if ds != "gsm8k" else EXPECTED[ds]*8,
                               "TODO","TODO","TODO","TODO","TODO","TODO","TODO","TODO"])

    print(f"E4: filled {filled} rows")
    return ws


def fill_e56_sheet(wb):
    """Rebuild E5-E6 Density Analysis sheet."""
    if "E5-E6 Density Analysis" in wb.sheetnames:
        idx = wb.sheetnames.index("E5-E6 Density Analysis")
        wb.remove(wb["E5-E6 Density Analysis"])
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet("E5-E6 Density Analysis", idx)

    headers = ["Model", "Dataset", "N\nCorrect", "N\nIncorrect",
               "Correct\nAvg Steps", "Incorrect\nAvg Steps", "\u0394 Steps\n(C\u2212I)",
               "Correct\nAvg \u03c1", "Incorrect\nAvg \u03c1", "\u0394 \u03c1\n(C\u2212I)",
               "Overall\n\u03c1 (Eq.2)", "Overall\nAccuracy(%)"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    filled = 0
    for model_list, family_label in [(QWEN_MODELS, None), (LLAMA_MODELS, "Llama Family")]:
        if family_label:
            ws.append([family_label])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        for model_san, model_disp, params_b in model_list:
            for ds in DATASETS:
                records = load_samples(model_san, ds)
                if records and len(records) >= EXPECTED[ds]:
                    stats = compute_e56_stats(records)
                    if stats:
                        n_c, n_i, c_s, i_s, d_s, c_r, i_r, d_r, o_r, o_a = stats
                        ws.append([model_disp, DS_DISPLAY[ds], n_c, n_i, c_s, i_s, d_s, c_r, i_r, d_r, o_r, o_a])
                        filled += 1
                    else:
                        ws.append([model_disp, DS_DISPLAY[ds]] + ["TODO"]*10)
                else:
                    ws.append([model_disp, DS_DISPLAY[ds]] + ["TODO"]*10)

    print(f"E5-E6: filled {filled} rows")
    return ws


def fill_status_matrix(wb):
    """Rebuild Status Matrix sheet."""
    if "Status Matrix" in wb.sheetnames:
        idx = wb.sheetnames.index("Status Matrix")
        wb.remove(wb["Status Matrix"])
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet("Status Matrix", idx)

    headers = ["Model", "GSM8K", "MATH-500", "AIME", "AMC", "Olympiad", "Total"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    total_done = defaultdict(int)
    total_all = defaultdict(int)

    for model_list, family_label in [(QWEN_MODELS, None), (LLAMA_MODELS, "Llama Family")]:
        if family_label:
            ws.append([family_label])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        for model_san, model_disp, params_b in model_list:
            short = STATUS_SHORT[model_san]
            row_data = [short]
            done_ct = 0
            for ds in DATASETS:
                records = load_samples(model_san, ds)
                complete = records is not None and len(records) >= EXPECTED[ds]
                status = "DONE" if complete else "TODO"
                row_data.append(status)
                total_all[ds] += 1
                if complete:
                    done_ct += 1
                    total_done[ds] += 1
            row_data.append(f"{done_ct}/5")
            ws.append(row_data)

    # Summary row
    total_models = len(ALL_MODELS)
    row_data = ["Done Count"]
    grand_done = 0
    grand_all = 0
    for ds in DATASETS:
        row_data.append(f"{total_done[ds]}/{total_all[ds]}")
        grand_done += total_done[ds]
        grand_all += total_all[ds]
    row_data.append(f"{grand_done}/{grand_all}")
    ws.append(row_data)
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    print(f"Status Matrix: {grand_done}/{grand_all} done")
    return ws


def generate_plots():
    """Generate E4 plots including reasoning density vs accuracy."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from scipy import stats

    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    # Collect per-model, per-dataset aggregated data
    model_data = {}  # (model_san, ds) -> (acc, avg_rho, avg_steps, params_b)
    for model_san, model_disp, params_b in ALL_MODELS:
        for ds in DATASETS:
            records = load_samples(model_san, ds)
            if records and len(records) >= EXPECTED[ds]:
                n_q, n_s, acc, std_acc, avg_s, std_s, avg_r, std_r, avg_t = compute_e4_stats(records)
                model_data[(model_san, ds)] = (acc, avg_r, avg_s, params_b, model_disp)

    # ── Plot 1: Model-level Reasoning Density vs Accuracy (per dataset) ──
    plt.rcParams.update({"font.size": 12, "font.family": "sans-serif"})

    fig, axes = plt.subplots(2, 3, figsize=(18, 11))
    axes_flat = axes.flatten()

    for di, ds in enumerate(DATASETS):
        ax = axes_flat[di]
        xs, ys, labels, sizes = [], [], [], []
        for model_san, model_disp, params_b in ALL_MODELS:
            key = (model_san, ds)
            if key in model_data:
                acc, avg_r, avg_s, pb, md = model_data[key]
                xs.append(avg_r)
                ys.append(acc)
                short = md.replace("-Instruct", "").replace("Qwen2.5-", "Q").replace("Llama-3.2-", "L3.2-").replace("Llama-3.1-", "L3.1-")
                labels.append(short)
                sizes.append(pb)

        xs, ys = np.array(xs), np.array(ys)
        # Color by family
        colors = ["#2563EB" if "Qwen" in l or "Q" in l else "#DC2626" for l in labels]
        marker_sizes = [max(40, min(200, s * 3)) for s in sizes]

        ax.scatter(xs, ys, c=colors, s=marker_sizes, alpha=0.8, edgecolors="white", linewidth=0.5, zorder=5)
        for i, label in enumerate(labels):
            ax.annotate(label, (xs[i], ys[i]), textcoords="offset points", xytext=(6, 4), fontsize=8)

        if len(xs) >= 3:
            slope, intercept, r_value, p_value, _ = stats.linregress(xs, ys)
            x_fit = np.linspace(xs.min() - 2, xs.max() + 2, 100)
            y_fit = slope * x_fit + intercept
            ax.plot(x_fit, y_fit, "--", color="gray", alpha=0.6, linewidth=1.2)
            rho_s, p_s = stats.spearmanr(xs, ys)
            ax.text(0.05, 0.05, f"Spearman \u03c1={rho_s:.2f} (p={p_s:.3f})\nPearson r={r_value:.2f}",
                    transform=ax.transAxes, fontsize=8, verticalalignment="bottom",
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="wheat", alpha=0.5))

        ax.set_xlabel("Avg Reasoning Density \u03c1 (tok/step)")
        ax.set_ylabel("Accuracy (%)")
        ax.set_title(f"{DS_DISPLAY[ds]}", fontweight="bold")
        ax.grid(alpha=0.3)

    # Panel 6: Combined across all datasets
    ax = axes_flat[5]
    xs_all, ys_all, labels_all, colors_all = [], [], [], []
    for model_san, model_disp, params_b in ALL_MODELS:
        rhos, accs = [], []
        for ds in DATASETS:
            key = (model_san, ds)
            if key in model_data:
                acc, avg_r, avg_s, pb, md = model_data[key]
                rhos.append(avg_r)
                accs.append(acc)
        if rhos:
            xs_all.append(np.mean(rhos))
            ys_all.append(np.mean(accs))
            short = model_disp.replace("-Instruct", "")
            labels_all.append(short)
            colors_all.append("#2563EB" if "Qwen" in model_disp else "#DC2626")

    xs_all, ys_all = np.array(xs_all), np.array(ys_all)
    ax.scatter(xs_all, ys_all, c=colors_all, s=100, alpha=0.8, edgecolors="white", zorder=5)
    for i, label in enumerate(labels_all):
        ax.annotate(label, (xs_all[i], ys_all[i]), textcoords="offset points", xytext=(6, 4), fontsize=8)
    if len(xs_all) >= 3:
        slope, intercept, r_value, p_value, _ = stats.linregress(xs_all, ys_all)
        x_fit = np.linspace(xs_all.min() - 2, xs_all.max() + 2, 100)
        y_fit = slope * x_fit + intercept
        ax.plot(x_fit, y_fit, "--", color="gray", alpha=0.6, linewidth=1.2)
        rho_s, p_s = stats.spearmanr(xs_all, ys_all)
        ax.text(0.05, 0.05, f"Spearman \u03c1={rho_s:.2f} (p={p_s:.3f})\nPearson r={r_value:.2f}",
                transform=ax.transAxes, fontsize=8, verticalalignment="bottom",
                bbox=dict(boxstyle="round,pad=0.3", facecolor="wheat", alpha=0.5))
    ax.set_xlabel("Avg Reasoning Density \u03c1 (tok/step)")
    ax.set_ylabel("Avg Accuracy (%)")
    ax.set_title("All Datasets (Averaged)", fontweight="bold")
    ax.grid(alpha=0.3)

    plt.suptitle("E4 Fig1: Reasoning Density vs Accuracy (Model-Level)", fontsize=14, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    p1 = PLOT_DIR / "e4_density_vs_accuracy_per_dataset.png"
    plt.savefig(p1, dpi=300, bbox_inches="tight")
    plt.close()
    print(f"Saved: {p1}")

    # ── Plot 2: Scaling plot (params vs accuracy, per dataset) ──
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    # Panel A: Qwen family scaling
    ax = axes[0]
    for ds in DATASETS:
        xs, ys = [], []
        for model_san, model_disp, params_b in QWEN_MODELS:
            key = (model_san, ds)
            if key in model_data:
                acc = model_data[key][0]
                xs.append(params_b)
                ys.append(acc)
        if xs:
            ax.plot(xs, ys, "-o", label=DS_DISPLAY[ds], markersize=6)
    ax.set_xscale("log")
    ax.set_xlabel("Model Size (B params)")
    ax.set_ylabel("Accuracy (%)")
    ax.set_title("Qwen2.5 Family Scaling", fontweight="bold")
    ax.legend(fontsize=9)
    ax.grid(alpha=0.3)

    # Panel B: Llama family scaling
    ax = axes[1]
    for ds in DATASETS:
        xs, ys = [], []
        for model_san, model_disp, params_b in LLAMA_MODELS:
            key = (model_san, ds)
            if key in model_data:
                acc = model_data[key][0]
                xs.append(params_b)
                ys.append(acc)
        if xs:
            ax.plot(xs, ys, "-s", label=DS_DISPLAY[ds], markersize=6)
    ax.set_xscale("log")
    ax.set_xlabel("Model Size (B params)")
    ax.set_ylabel("Accuracy (%)")
    ax.set_title("Llama Family Scaling", fontweight="bold")
    ax.legend(fontsize=9)
    ax.grid(alpha=0.3)

    plt.tight_layout()
    p2 = PLOT_DIR / "e4_scaling_params_vs_accuracy.png"
    plt.savefig(p2, dpi=300, bbox_inches="tight")
    plt.close()
    print(f"Saved: {p2}")

    # ── Plot 3: Reasoning Density vs Accuracy (reviewer request) ──
    # This is the key plot: for each model, plot (overall rho, accuracy) across datasets
    fig, ax = plt.subplots(1, 1, figsize=(10, 7))

    # Collect all (rho, acc) points colored by dataset, sized by model
    ds_colors = {"gsm8k": "#2563EB", "math500": "#059669", "aime": "#DC2626", "amc": "#7C3AED", "olympiad": "#D97706"}
    ds_markers = {"gsm8k": "o", "math500": "s", "aime": "D", "amc": "^", "olympiad": "v"}

    for ds in DATASETS:
        xs, ys, labels = [], [], []
        for model_san, model_disp, params_b in ALL_MODELS:
            key = (model_san, ds)
            if key in model_data:
                acc, avg_r, avg_s, pb, md = model_data[key]
                xs.append(avg_r)
                ys.append(acc)
                labels.append(md.replace("-Instruct", ""))
        if xs:
            ax.scatter(xs, ys, c=ds_colors[ds], marker=ds_markers[ds], s=80, alpha=0.8,
                       label=DS_DISPLAY[ds], edgecolors="white", linewidth=0.5, zorder=5)

    # Overall regression line
    all_x, all_y = [], []
    for key, (acc, avg_r, avg_s, pb, md) in model_data.items():
        all_x.append(avg_r)
        all_y.append(acc)
    all_x, all_y = np.array(all_x), np.array(all_y)
    if len(all_x) >= 3:
        slope, intercept, r_value, p_value, _ = stats.linregress(all_x, all_y)
        x_fit = np.linspace(all_x.min() - 5, all_x.max() + 5, 100)
        y_fit = slope * x_fit + intercept
        ax.plot(x_fit, y_fit, "--", color="gray", alpha=0.5, linewidth=1.5, label=f"Linear fit (r={r_value:.2f})")
        rho_s, p_s = stats.spearmanr(all_x, all_y)
        ax.text(0.02, 0.98, f"Spearman \u03c1 = {rho_s:.3f} (p = {p_s:.1e})\nPearson r = {r_value:.3f} (p = {p_value:.1e})\nN = {len(all_x)} points",
                transform=ax.transAxes, fontsize=10, verticalalignment="top",
                bbox=dict(boxstyle="round,pad=0.4", facecolor="lightyellow", alpha=0.8))

    ax.set_xlabel("Reasoning Density \u03c1 (tokens/step)", fontsize=12)
    ax.set_ylabel("Accuracy (%)", fontsize=12)
    ax.set_title("Reasoning Density (Eq. 2) vs Accuracy\nAcross Models and Datasets", fontsize=13, fontweight="bold")
    ax.legend(fontsize=10, loc="lower right")
    ax.grid(alpha=0.3)

    plt.tight_layout()
    p3 = PLOT_DIR / "e4_reasoning_density_vs_accuracy_reviewer.png"
    plt.savefig(p3, dpi=400, bbox_inches="tight")
    plt.close()
    print(f"Saved: {p3}")

    # ── Plot 4: Within-model density analysis (correct vs incorrect) ──
    fig, axes = plt.subplots(2, 3, figsize=(18, 11))
    axes_flat = axes.flatten()

    for di, ds in enumerate(DATASETS):
        ax = axes_flat[di]
        model_labels = []
        correct_rhos = []
        incorrect_rhos = []
        for model_san, model_disp, params_b in QWEN_MODELS:
            records = load_samples(model_san, ds)
            if records and len(records) >= EXPECTED[ds]:
                c_r = [s["density_rho"] for rec in records for s in rec["samples"] if s["correct"]]
                i_r = [s["density_rho"] for rec in records for s in rec["samples"] if not s["correct"]]
                if c_r and i_r:
                    model_labels.append(model_disp.replace("Qwen2.5-","").replace("-Instruct",""))
                    correct_rhos.append(c_r)
                    incorrect_rhos.append(i_r)

        if model_labels:
            positions = list(range(len(model_labels)))
            width = 0.35
            bp1 = ax.boxplot(correct_rhos, positions=[p - width/2 for p in positions],
                            widths=width*0.8, patch_artist=True, showfliers=False,
                            boxprops=dict(facecolor="#BFDBFE", edgecolor="#2563EB"),
                            medianprops=dict(color="#1E40AF", linewidth=2))
            bp2 = ax.boxplot(incorrect_rhos, positions=[p + width/2 for p in positions],
                            widths=width*0.8, patch_artist=True, showfliers=False,
                            boxprops=dict(facecolor="#FECACA", edgecolor="#DC2626"),
                            medianprops=dict(color="#991B1B", linewidth=2))
            ax.set_xticks(positions)
            ax.set_xticklabels(model_labels, fontsize=9)
            ax.legend([bp1["boxes"][0], bp2["boxes"][0]], ["Correct", "Incorrect"], fontsize=9)
        ax.set_ylabel("\u03c1 (tok/step)")
        ax.set_title(f"{DS_DISPLAY[ds]}: Correct vs Incorrect \u03c1", fontweight="bold")
        ax.grid(axis="y", alpha=0.3)

    # Panel 6: empty or summary
    ax = axes_flat[5]
    ax.axis("off")
    ax.text(0.5, 0.5, "Qwen2.5 Family\nCorrect vs Incorrect\nReasoning Density",
            transform=ax.transAxes, ha="center", va="center", fontsize=14, style="italic")

    plt.suptitle("E5-E6: Within-Model Density Analysis (Correct vs Incorrect)", fontsize=14, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    p4 = PLOT_DIR / "e56_correct_vs_incorrect_density.png"
    plt.savefig(p4, dpi=300, bbox_inches="tight")
    plt.close()
    print(f"Saved: {p4}")


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)

    print("=== Filling E4 Fig1 Scaling ===")
    fill_e4_sheet(wb)

    print("\n=== Filling E5-E6 Density Analysis ===")
    fill_e56_sheet(wb)

    print("\n=== Filling Status Matrix ===")
    fill_status_matrix(wb)

    wb.save(XLSX_PATH)
    print(f"\nSaved: {XLSX_PATH}")

    print("\n=== Generating Plots ===")
    generate_plots()

    print("\nDone!")


if __name__ == "__main__":
    main()
