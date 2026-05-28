#!/usr/bin/env python3
"""
Collect E0-E3 sweep results and update rebuttal_experiment_data.xlsx.

Reads all completed sweep eval results, finds best lambda per experiment,
and writes to the "E0-E3 Controls" sheet.

Usage:
    python update_e0e3_excel.py                  # one-shot update
    python update_e0e3_excel.py --watch 120      # update every 120s until sweep done
"""

import argparse
import glob
import json
import os
import sys
import time
from datetime import datetime

BASE = "/common/users/sl2148/Public/yang_ouyang/projects/fact-enhancement"
CONTROL_DIR = os.path.join(BASE, "control_experiments", "Qwen_Qwen2.5-3B-Instruct")
SWEEP_EVAL = os.path.join(CONTROL_DIR, "sweep_eval")
STATUS_PATH = os.path.join(CONTROL_DIR, "sweep_status.json")
XLSX_PATH = os.path.join(BASE, "documents", "rebuttal_experiment_data.xlsx")

EXPERIMENT_LABELS = {
    "e0": "Random Paraphrase Steering",
    "e1": "Random Step Compression Steering",
    "e2": "Dense-but-Incorrect Steering",
    "e3": "Rule-Based Rewriting Steering",
    "e3.2": "GPT-5-mini Rewriting Steering",
}

EXPERIMENT_IDS = {
    "e0": "E0",
    "e1": "E1",
    "e2": "E2",
    "e3": "E3",
    "e3.2": "E3.2",
}


def get_accuracy(results_dict, task="gsm8k_cot_zeroshot_unified"):
    task_results = results_dict.get(task, {})
    for key_prefix in ["exact_match"]:
        for suffix in [",flexible-extract", ",none"]:
            key = f"{key_prefix}{suffix}"
            if key in task_results:
                acc = task_results[key]
                se_key = f"{key_prefix}_stderr{suffix}"
                se = task_results.get(se_key, 0)
                return acc, se
    return None, None


def collect_sweep_results():
    """Scan sweep_eval dirs and collect accuracy for each (experiment, lambda)."""
    results = {}  # {exp_name: [(lambda, acc, se, outdir), ...]}

    if not os.path.exists(SWEEP_EVAL):
        return results

    for entry in sorted(os.listdir(SWEEP_EVAL)):
        entry_path = os.path.join(SWEEP_EVAL, entry)
        if not os.path.isdir(entry_path):
            continue

        # Parse experiment name and lambda from dir name like e0_L6_lam4p0
        parts = entry.split("_")
        if len(parts) < 3:
            continue

        # Handle e3.2 -> "e3.2"
        exp_name = parts[0]
        if exp_name == "e3" and len(parts) >= 4 and parts[1] == "2":
            exp_name = "e3.2"
            lam_part = parts[3] if len(parts) > 3 else ""
            layer_part = parts[2]
        elif exp_name == "e3.2":
            layer_part = parts[1]
            lam_part = parts[2] if len(parts) > 2 else ""
        else:
            layer_part = parts[1]
            lam_part = parts[2] if len(parts) > 2 else ""

        # Find results JSON
        result_files = glob.glob(
            os.path.join(entry_path, "**", "results_*.json"), recursive=True
        )
        if not result_files:
            continue

        try:
            with open(result_files[0]) as f:
                data = json.load(f)
            acc, se = get_accuracy(data.get("results", {}))
            if acc is None:
                continue

            # Parse lambda from model_args
            model_args = data.get("config", {}).get("model_args", "")
            lam = None
            layer = None
            for arg in model_args.split(","):
                if "steer_lambda=" in arg:
                    lam = float(arg.split("=")[1])
                if "steer_layer=" in arg:
                    layer = int(arg.split("=")[1])

            if lam is None:
                continue

            results.setdefault(exp_name, []).append(
                {"lambda": lam, "layer": layer, "acc": acc, "se": se, "dir": entry}
            )
        except Exception as e:
            print(f"  [WARN] Error reading {entry}: {e}")

    return results


def find_best_per_experiment(results):
    """For each experiment, find the lambda with highest accuracy."""
    best = {}
    for exp_name, entries in results.items():
        if not entries:
            continue
        top = max(entries, key=lambda x: x["acc"])
        best[exp_name] = top
        # Also find baseline (lambda=0)
        baseline = [e for e in entries if abs(e["lambda"]) < 0.01]
        if baseline:
            best[exp_name]["baseline_acc"] = baseline[0]["acc"]
    return best


def update_excel(best_results, all_results):
    """Update the E0-E3 Controls sheet in the Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed, skipping Excel update")
        return False

    if not os.path.exists(XLSX_PATH):
        print(f"Excel file not found: {XLSX_PATH}")
        return False

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["E0-E3 Controls"]

    # Find the row structure - scan for experiment IDs in column A
    exp_rows = {}
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        if cell_a and cell_b:
            cell_a_str = str(cell_a).strip()
            cell_b_str = str(cell_b).strip()
            for exp_name, exp_id in EXPERIMENT_IDS.items():
                if cell_a_str == exp_id:
                    exp_rows.setdefault(exp_name, []).append(row_idx)

    # Column mapping (1-indexed):
    # A=Exp, B=Method, C=Model, D=Layer, E=λ, F=GSM8K Acc(%), G=GSM8K Δ vs Base,
    # H=GSM8K Δ vs Dense, I=MATH-500, J=AIME, K=AMC, L=Avg Steps, M=Avg ρ, N=Notes
    COL_EXP = 1
    COL_METHOD = 2
    COL_MODEL = 3
    COL_LAYER = 4
    COL_LAMBDA = 5
    COL_GSM8K_ACC = 6
    COL_GSM8K_DELTA_BASE = 7
    COL_GSM8K_DELTA_DENSE = 8
    COL_NOTES = 14

    BASELINE_ACC = 83.8  # from the sheet
    DENSE_BEST_ACC = 85.9  # from the sheet

    updated_count = 0

    for exp_name, best in best_results.items():
        if exp_name not in exp_rows:
            print(f"  [SKIP] No row found for {exp_name} in Excel")
            continue

        # Use first row for this experiment (Qwen2.5-3B-Instruct)
        row = exp_rows[exp_name][0]

        acc_pct = round(best["acc"] * 100, 1)
        delta_base = round(acc_pct - BASELINE_ACC, 1)
        delta_dense = round(acc_pct - DENSE_BEST_ACC, 1)

        ws.cell(row=row, column=COL_MODEL, value="Qwen2.5-3B-Instruct")
        ws.cell(row=row, column=COL_LAYER, value=best.get("layer", 6))
        ws.cell(row=row, column=COL_LAMBDA, value=best["lambda"])
        ws.cell(row=row, column=COL_GSM8K_ACC, value=acc_pct)
        ws.cell(row=row, column=COL_GSM8K_DELTA_BASE,
                value=f"+{delta_base}" if delta_base > 0 else str(delta_base))
        ws.cell(row=row, column=COL_GSM8K_DELTA_DENSE,
                value=f"+{delta_dense}" if delta_dense > 0 else str(delta_dense))

        # Build notes with all lambda results
        exp_entries = sorted(all_results.get(exp_name, []), key=lambda x: x["lambda"])
        n_done = len(exp_entries)
        note = f"Best λ={best['lambda']}, {n_done}/21 lambdas done"
        if best.get("se"):
            note += f", SE={best['se']*100:.2f}%"
        ws.cell(row=row, column=COL_NOTES, value=note)

        updated_count += 1
        print(f"  Updated {EXPERIMENT_IDS[exp_name]} row {row}: "
              f"acc={acc_pct}% (λ={best['lambda']}, Δbase={delta_base:+.1f}, Δdense={delta_dense:+.1f})")

    if updated_count > 0:
        wb.save(XLSX_PATH)
        print(f"\n  Saved {XLSX_PATH} ({updated_count} experiments updated)")
    else:
        print("  No updates to write.")

    return updated_count > 0


def print_summary(all_results, best_results):
    """Print a summary table of all results."""
    print(f"\n{'='*80}")
    print(f"E0-E3 Sweep Results Summary  ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    print(f"{'='*80}")
    print(f"{'Exp':<6} {'Method':<35} {'Best λ':>7} {'Acc%':>7} {'ΔBase':>7} {'ΔDense':>7} {'N done':>7}")
    print("-" * 80)

    BASELINE_ACC = 83.8
    DENSE_BEST_ACC = 85.9

    for exp_name in ["e0", "e1", "e2", "e3", "e3.2"]:
        label = EXPERIMENT_LABELS.get(exp_name, exp_name)
        entries = all_results.get(exp_name, [])
        best = best_results.get(exp_name)

        if best:
            acc_pct = best["acc"] * 100
            delta_base = acc_pct - BASELINE_ACC
            delta_dense = acc_pct - DENSE_BEST_ACC
            print(f"{EXPERIMENT_IDS[exp_name]:<6} {label:<35} {best['lambda']:>7.1f} "
                  f"{acc_pct:>6.1f}% {delta_base:>+6.1f} {delta_dense:>+6.1f} "
                  f"{len(entries):>3}/21")
        else:
            print(f"{EXPERIMENT_IDS[exp_name]:<6} {label:<35} {'—':>7} {'—':>7} {'—':>7} {'—':>7} "
                  f"{len(entries):>3}/21")

    # Also print full lambda sweep for each experiment
    for exp_name in ["e0", "e1", "e2", "e3", "e3.2"]:
        entries = sorted(all_results.get(exp_name, []), key=lambda x: x["lambda"])
        if not entries:
            continue
        print(f"\n  {EXPERIMENT_IDS[exp_name]} lambda sweep ({len(entries)} points):")
        for e in entries:
            marker = " ★" if best_results.get(exp_name, {}).get("lambda") == e["lambda"] else ""
            print(f"    λ={e['lambda']:>6.1f}  acc={e['acc']*100:>6.2f}%{marker}")


def get_sweep_status():
    """Read sweep status to check completion."""
    if not os.path.exists(STATUS_PATH):
        return {}
    with open(STATUS_PATH) as f:
        data = json.load(f)
    jobs = data.get("jobs", {})
    by_status = {}
    for v in jobs.values():
        s = v.get("status", "unknown")
        by_status[s] = by_status.get(s, 0) + 1
    return by_status


def main():
    parser = argparse.ArgumentParser(description="Update E0-E3 Excel with sweep results")
    parser.add_argument("--watch", type=int, default=0,
                        help="Re-check every N seconds until sweep done (0=one-shot)")
    parser.add_argument("--no-excel", action="store_true",
                        help="Print results only, don't update Excel")
    args = parser.parse_args()

    while True:
        status = get_sweep_status()
        total = sum(status.values())
        done = status.get("done", 0)
        running = status.get("running", 0)
        pending = status.get("pending", 0)
        print(f"\nSweep status: {done}/{total} done, {running} running, {pending} pending")

        all_results = collect_sweep_results()
        best_results = find_best_per_experiment(all_results)
        print_summary(all_results, best_results)

        if not args.no_excel:
            print("\nUpdating Excel...")
            update_excel(best_results, all_results)

        if args.watch <= 0:
            break

        if pending == 0 and running == 0:
            print("\nSweep complete! Final update done.")
            break

        print(f"\nWaiting {args.watch}s before next check...")
        time.sleep(args.watch)


if __name__ == "__main__":
    main()
