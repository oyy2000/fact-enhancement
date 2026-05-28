#!/usr/bin/env python3
"""Populate documents/rebuttal_experiment_data.xlsx → sheet 'E8 Calibration Size'."""
import glob
import json
import os
import sys

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from transformers import AutoTokenizer

_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)
from utils import split_double_newline  # noqa: E402

XLSX = "/common/users/sl2148/Public/yang_ouyang/projects/fact-enhancement/documents/rebuttal_experiment_data.xlsx"
BASE = "/common/users/sl2148/Public/yang_ouyang/projects/fact-enhancement"
MODEL = "Qwen2.5-3B-Instruct"
HF_MODEL = "Qwen/Qwen2.5-3B-Instruct"

STATUS_GPT = os.path.join(
    BASE,
    "calibration_ablation/Qwen_Qwen2.5-3B-Instruct/GPT_REWRITE/formal_L6_limit1000_status.json",
)
STATUS_LM = os.path.join(
    BASE,
    "calibration_ablation/Qwen_Qwen2.5-3B-Instruct/"
    "LARGE_MODEL_Qwen_Qwen2.5-7B-Instruct/formal_L6_limit1000_status.json",
)

# Formal eval (limit 1000, L=6, λ = pilot L6-best); Pilot EM = limit 100 L6 pilot
GPT_ROWS = [
    (1, 6, -3.5, 0.840, 0.840),
    (5, 6, -4.5, 0.830, 0.841),
    (10, 6, -1.0, 0.830, 0.836),
    (25, 6, 5.0, 0.850, 0.842),
    (50, 6, 5.0, 0.840, 0.845),
    (100, 6, -1.5, 0.830, 0.836),
    (200, 6, 3.5, 0.830, 0.839),
]
LM_ROWS = [
    (1, 6, 0.4, 0.850, 0.842),
    (5, 6, 0.6, 0.850, 0.852),
    (10, 6, 0.5, 0.850, 0.843),
    (25, 6, 0.4, 0.850, 0.858),
    (50, 6, 0.4, 0.850, 0.852),
    (100, 6, 0.4, 0.850, 0.851),
    (200, 6, 0.6, 0.850, 0.825),
]

# Short per-row notes (Chinese)
ROW_NOTES = {
    1: "仅 1 对正反例：方向估计方差极大，极易受该样本噪声与难度支配；pilot(n=100) 与正式(n=1000) 可比性弱。",
    5: "样本仍极少，均值向量不稳定；对「从尾部 doc_id 选取 EM=1」策略更敏感。",
    10: "开始有平滑但仍偏少；λ 在 pilot 上选优后，换更大 eval 集方差仍可见。",
    25: "中等规模：方向与范数更稳，GPT_REWRITE 上 λ=5.0 在正式集表现较好。",
    50: "论文默认校准规模；作为 Δ 对照锚点。",
    100: "更大 N 使向量更接近「平均改写差分」；可能略保守或峰值 λ 偏移（仍依赖 pilot 网格）。",
    200: "平均化最强：信号可能被稀释；GPT_REWRITE 数据上 N=200 实际仅用 170 对（EM=1 不足 200），与名义 N 不一致。",
}

# Block: same N, both modes — extra interpretation
N_INTERPRETATION = {
    1: (
        "N=1：理论上估计误差最大；两模式正式 EM 均接近 pilot，但不宜外推为稳定最优。"
    ),
    5: (
        "N=5：小样本下 steering 方向仍带随机性；GPT 与 LM 来源不同（GPT 改写 vs 7B 配对），"
        "最优 λ 尺度不同属预期（范数量级差异）。"
    ),
    10: (
        "N=10：性能通常介于极小 N 与论文默认之间；可作「低成本复现」参考点。"
    ),
    25: (
        "N=25：LARGE_MODEL 在正式集上峰值最高(85.8%)，说明该规模下 7B-配对向量与 L6、λ≈0.4–0.6 更匹配；"
        "仍依赖 pilot 在 limit=100 上选 λ 带来的评估噪声。"
    ),
    50: (
        "N=50：论文主设置；两模式正式 EM 分别为 84.5% / 85.2%，作为其它 N 的 Δ 基准。"
    ),
    100: (
        "N=100：向量更平滑；LARGE_MODEL 仍保持 ~85.1%，GPT_REWRITE 略低于 N=50 峰值，"
        "可能因 λ 在 pilot 上非全局最优或校准与 eval 分布轻微错配。"
    ),
    200: (
        "N=200：GPT_REWRITE 仅 170 对有效，名义 200 与实现不一致，解释需谨慎。"
        "LARGE_MODEL 正式集下降至 82.5%：可能原因包括——(1) pilot 上最优 λ=0.6 在 n=1000 上不再最优；"
        "(2) 更大 N 平均化削弱与目标任务对齐的「尖锐」方向；(3) 与 limit=1000 评估子集组合的随机波动。"
        "建议对照 stderr 或重复 seed / 微调 λ 验证。"
    ),
}


def compute_steps_rho_from_jsonl(jsonl_path, tokenizer):
    """Match E5–E6 / utils.process_single_file: double-newline steps, mean tok/step per sample."""
    steps_list = []
    rho_list = []
    with open(jsonl_path, encoding="utf-8") as f:
        for line in f:
            d = json.loads(line)
            if d.get("filter") == "strict-match":
                continue
            try:
                cot = d["resps"][0][0].strip()
            except (KeyError, IndexError, TypeError, AttributeError):
                continue
            steps_dbl = split_double_newline(cot)
            tokens_dbl = [
                len(tokenizer.encode(s, add_special_tokens=False)) for s in steps_dbl
            ]
            steps_list.append(len(steps_dbl))
            rho_list.append(float(np.mean(tokens_dbl)) if tokens_dbl else 0.0)
    if not steps_list:
        return None, None
    return (
        round(float(np.mean(steps_list)), 2),
        round(float(np.mean(rho_list)), 2),
    )


def collect_steps_rho_by_n(status_path, tokenizer):
    """calib_size -> (avg_steps, avg_rho) from done jobs' samples_*.jsonl."""
    out = {}
    if not os.path.isfile(status_path):
        print(f"[WARN] Missing status: {status_path}")
        return out
    with open(status_path, encoding="utf-8") as f:
        data = json.load(f)
    for rec in data.get("jobs", {}).values():
        if not isinstance(rec, dict):
            continue
        if rec.get("status") != "done" or rec.get("returncode") != 0:
            continue
        n = rec.get("calib_size")
        od = rec.get("outdir")
        if n is None or not od or not os.path.isdir(od):
            continue
        matches = sorted(
            glob.glob(os.path.join(od, "**", "samples_*.jsonl"), recursive=True)
        )
        if not matches:
            continue
        jsonl_path = matches[-1]
        ast, arh = compute_steps_rho_from_jsonl(jsonl_path, tokenizer)
        if ast is not None:
            out[int(n)] = (ast, arh)
    return out


def short_note(mode: str, n: int, formal: float) -> str:
    base = ROW_NOTES.get(n, "")
    if mode == "LARGE_MODEL" and n == 200:
        return (
            base
            + " LARGE_MODEL：正式集明显低于 pilot，优先检查 λ 迁移与平均化效应，而非仅归因样本数。"
        )
    return base


def main():
    wb = load_workbook(XLSX)
    name = "E8 Calibration Size"
    idx = wb.sheetnames.index(name)
    wb.remove(wb[name])
    ws = wb.create_sheet(name, idx)

    title = (
        "E8 Calibration Size | 被测模型：Qwen2.5-3B-Instruct | GSM8K | "
        "GPT_REWRITE：GPT 改写配对；LARGE_MODEL：Qwen2.5-7B-Instruct 配对改写。\n"
        "Layer 固定为 6；λ 为 pilot（limit=100）上第 6 层 flexible-extract EM 最大者；"
        "正式结果为 limit=1000、exact_match flexible-extract。\n"
        "Avg Steps / Avg ρ：与项目 E5–E6 一致——对模型输出按双换行分步，"
        "用 Qwen2.5-3B tokenizer 计每步 token 数；ρ 为每题「步内平均 token」再在题上取平均。"
    )
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws["A1"].font = Font(bold=True, size=11)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 56

    print("Loading tokenizer & scanning formal eval jsonl (may take ~1–2 min)...")
    tokenizer = AutoTokenizer.from_pretrained(HF_MODEL, trust_remote_code=True)
    metrics_gpt = collect_steps_rho_by_n(STATUS_GPT, tokenizer)
    metrics_lm = collect_steps_rho_by_n(STATUS_LM, tokenizer)
    print("GPT_REWRITE steps/ρ by N:", metrics_gpt)
    print("LARGE_MODEL steps/ρ by N:", metrics_lm)

    headers = [
        "Steering 来源",
        "N (pairs)",
        "Model",
        "Layer",
        "λ (L6 best, pilot)",
        "Pilot EM\n(L6, n=100)",
        "Formal EM\n(n=1000)",
        "GSM8K Acc (%)\n(formal)",
        "Δ vs N=50\n(pp, 同模式)",
        "Avg Steps",
        "Avg ρ",
        "本行简注",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def emit_rows(mode: str, rows, metrics_map):
        ref50 = next(r for r in rows if r[0] == 50)[4]
        for n, layer, lam, pilot, formal in rows:
            acc_pct = round(formal * 100, 2)
            ref_pct = ref50 * 100
            delta_pp = round(acc_pct - ref_pct, 2)
            pair = metrics_map.get(n)
            if pair:
                avg_steps, avg_rho = pair
            else:
                avg_steps, avg_rho = "—", "—"
            ws.append(
                [
                    mode,
                    n,
                    MODEL,
                    layer,
                    lam,
                    round(pilot, 4),
                    round(formal, 4),
                    acc_pct,
                    delta_pp,
                    avg_steps,
                    avg_rho,
                    short_note(mode, n, formal),
                ]
            )

    emit_rows("GPT_REWRITE", GPT_ROWS, metrics_gpt)
    emit_rows("LARGE_MODEL", LM_ROWS, metrics_lm)

    start_block = ws.max_row + 2
    ws.cell(row=start_block, column=1, value="各 N 档机制性说明（与 steering 来源无关的共性）")
    ws.cell(row=start_block, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(
        start_row=start_block, start_column=1, end_row=start_block, end_column=12
    )

    r = start_block + 1
    for n in sorted(N_INTERPRETATION.keys()):
        ws.cell(row=r, column=1, value=f"N = {n}")
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
        cell = ws.cell(row=r, column=2, value=N_INTERPRETATION[n])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    widths = [22, 10, 22, 8, 14, 14, 14, 14, 14, 12, 10, 56]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=3, max_row=2 + len(GPT_ROWS) + len(LM_ROWS)):
        for cell in row:
            if cell.column == 12:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(XLSX)
    print(f"Updated: {XLSX} sheet '{name}'")


if __name__ == "__main__":
    main()
