#!/usr/bin/env python3
"""
Watchdog script that:
1. Monitors BBH experiments (already running on GPUs 0,1,2)
2. Launches MuSiQue experiments when GPUs 3,4,5 are available
3. Updates the rebuttal_experiment_data.xlsx with all E12 results
"""
import subprocess, os, sys, time, json, glob

BASE = "/common/users/sl2148/Public/yang_ouyang/projects/fact-enhancement"
PYTHON = "/common/users/sl2148/anaconda3/envs/fact_yang/bin/python"
OUTBASE = f"{BASE}/exps/steer_runs_e11_e12"
LOGDIR = f"{OUTBASE}/overnight_logs"
os.makedirs(LOGDIR, exist_ok=True)

hf_token = ""
token_path = f"{BASE}/new_exps/.cache/huggingface/token"
if os.path.exists(token_path):
    with open(token_path) as f:
        hf_token = f.read().strip()

MODEL = "Qwen/Qwen2.5-3B-Instruct"
GPT_VEC = f"{BASE}/exps/gpt_rewrites_unified_new/Qwen_Qwen2.5-3B-Instruct/vectors_50_old/Qwen_Qwen2.5-3B-Instruct_applied/steering_vector.pt"
LM_VEC = f"{BASE}/exps/large_model_rewrites_unified_new/Qwen_Qwen2.5-3B-Instruct/vectors_50_paired_Qwen_Qwen2.5-7B-Instruct/Qwen_Qwen2.5-3B-Instruct_applied/steering_vector.pt"

def log(msg):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)

def find_results(mode, task):
    safe_model = MODEL.replace("/", "_")
    pattern = f"{OUTBASE}/{mode}/{safe_model}/{task}/**/results_*.json"
    files = glob.glob(pattern, recursive=True)
    if not files:
        pattern2 = f"{OUTBASE}/{mode}/Qwen__Qwen2.5-3B-Instruct/{task}/**/results_*.json"
        files = glob.glob(pattern2, recursive=True)
    return files

def get_score(results_file, task):
    with open(results_file) as f:
        data = json.load(f)
    results = data.get("results", {})
    if task in results:
        for key, val in results[task].items():
            if "stderr" not in key and key != "alias":
                return key, val
    return None, None

def launch_lm_eval(gpu, task, mode, layer, lam, vec, batch_size=1, max_gen_toks=512):
    safe_model = MODEL.replace("/", "_")
    lam_tag = f"{lam:.6f}".rstrip("0").rstrip(".")
    outdir = f"{OUTBASE}/{mode}/{safe_model}/{task}/L{layer}/lam_{lam_tag}"
    logfile = f"{LOGDIR}/{mode}_{task}_L{layer}_lam{lam_tag}.log"

    existing = glob.glob(f"{outdir}/**/results_*.json", recursive=True)
    if existing:
        log(f"[SKIP] {mode} {task} - already done")
        return None

    os.makedirs(outdir, exist_ok=True)

    cmd = [
        PYTHON, "-m", "lm_eval",
        "--model", "steer_hf",
        "--model_args", f"pretrained={MODEL},dtype=float16,steer_layer={layer},steer_lambda={lam},steer_vec_path={vec}",
        "--tasks", task,
        "--device", "cuda:0",
        "--num_fewshot", "0",
        "--batch_size", str(batch_size),
        "--gen_kwargs", f"do_sample=false,temperature=0,max_gen_toks={max_gen_toks}",
        "--output_path", outdir,
        "--log_samples",
        "--trust_remote_code",
        "--apply_chat_template",
    ]

    env = os.environ.copy()
    env["CUDA_VISIBLE_DEVICES"] = str(gpu)
    env["HF_TOKEN"] = hf_token
    env["HUGGING_FACE_HUB_TOKEN"] = hf_token

    log(f"[START] GPU={gpu} {mode} {task} L{layer} lam={lam} batch={batch_size}")
    log_fh = open(logfile, "w")
    p = subprocess.Popen(cmd, stdout=log_fh, stderr=subprocess.STDOUT, env=env, cwd=BASE)
    return {"proc": p, "log_fh": log_fh, "mode": mode, "task": task, "gpu": gpu}

def wait_for_tasks(tasks):
    remaining = list(tasks)
    while remaining:
        for t in remaining[:]:
            rc = t["proc"].poll()
            if rc is not None:
                t["log_fh"].close()
                status = "DONE" if rc == 0 else "FAIL"
                log(f"[{status}] {t['mode']} {t['task']} GPU={t['gpu']} (rc={rc})")
                remaining.remove(t)
        if remaining:
            time.sleep(30)
    return

def update_xlsx():
    """Update the Excel file with all E12 results"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        log("openpyxl not available, skipping XLSX update")
        return

    xlsx_path = f"{BASE}/documents/rebuttal_experiment_data.xlsx"
    if not os.path.exists(xlsx_path):
        log(f"XLSX not found at {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    if "E12 Non-Math Reasoning" not in wb.sheetnames:
        log("E12 sheet not found, creating it")
        ws = wb.create_sheet("E12 Non-Math Reasoning")
    else:
        ws = wb["E12 Non-Math Reasoning"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

    headers = [
        "Steering Method", "Model", "Layer", "Lambda",
        "", "BBH CoT (acc_norm)", "HotpotQA (F1)", "MuSiQue (F1)",
        "", "BBH Delta", "HotpotQA Delta", "MuSiQue Delta",
        "Status", "Notes"
    ]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = header_font

    modes = [
        {"name": "Baseline (No Steering)", "key": "BASELINE", "layer": "—", "lam": "—",
         "actual_layer": 0, "actual_lam": 0.0},
        {"name": "DenseSteer (GPT_REWRITE)", "key": "GPT_REWRITE", "layer": 6, "lam": 4.0,
         "actual_layer": 6, "actual_lam": 4.0},
        {"name": "InFamilySteer (LARGE_MODEL)", "key": "LARGE_MODEL", "layer": 6, "lam": 0.45,
         "actual_layer": 6, "actual_lam": 0.45},
    ]

    baseline_scores = {}

    for row_idx, mode in enumerate(modes, 2):
        ws.cell(row=row_idx, column=1, value=mode["name"])
        ws.cell(row=row_idx, column=2, value="Qwen2.5-3B-Instruct")
        ws.cell(row=row_idx, column=3, value=mode["layer"])
        ws.cell(row=row_idx, column=4, value=mode["lam"])

        for task, col, delta_col, task_key in [
            ("bbh_cot_zeroshot", 6, 10, "bbh"),
            ("longbench_hotpotqa", 7, 11, "hotpotqa"),
            ("longbench_musique", 8, 12, "musique"),
        ]:
            results_files = find_results(mode["key"], task)
            if results_files:
                _, score = get_score(results_files[0], task)
                if score is not None:
                    score_pct = score * 100 if score < 1 else score
                    ws.cell(row=row_idx, column=col, value=round(score_pct, 2))
                    if mode["key"] == "BASELINE":
                        baseline_scores[task_key] = score_pct
                    elif task_key in baseline_scores:
                        delta = score_pct - baseline_scores[task_key]
                        ws.cell(row=row_idx, column=delta_col, value=round(delta, 2))
                        color = "00AA00" if delta >= 0 else "CC0000"
                        ws.cell(row=row_idx, column=delta_col).font = Font(color=color)
                else:
                    ws.cell(row=row_idx, column=col, value="ERROR")
            else:
                ws.cell(row=row_idx, column=col, value="—")

        all_done = all(find_results(mode["key"], t) for t in ["bbh_cot_zeroshot", "longbench_hotpotqa", "longbench_musique"])
        ws.cell(row=row_idx, column=13, value="DONE" if all_done else "PARTIAL")

    wb.save(xlsx_path)
    log(f"Updated {xlsx_path} with E12 results")

if __name__ == "__main__":
    log("=" * 60)
    log("Watchdog Overnight Script Started")
    log("=" * 60)

    # Step 1: Check what's already done
    all_tasks = []
    for task in ["bbh_cot_zeroshot", "longbench_hotpotqa", "longbench_musique"]:
        for mode_key in ["BASELINE", "GPT_REWRITE", "LARGE_MODEL"]:
            r = find_results(mode_key, task)
            status = "DONE" if r else "TODO"
            log(f"  {mode_key:15s} {task:25s} -> {status}")
            all_tasks.append((mode_key, task, bool(r)))

    # Step 2: Launch MuSiQue if not done
    musique_tasks = []
    musique_configs = [
        {"gpu": 3, "mode": "BASELINE",    "layer": 0, "lam": 0.0,  "vec": GPT_VEC},
        {"gpu": 4, "mode": "GPT_REWRITE", "layer": 6, "lam": 4.0,  "vec": GPT_VEC},
        {"gpu": 5, "mode": "LARGE_MODEL", "layer": 6, "lam": 0.45, "vec": LM_VEC},
    ]

    for cfg in musique_configs:
        t = launch_lm_eval(cfg["gpu"], "longbench_musique", cfg["mode"],
                          cfg["layer"], cfg["lam"], cfg["vec"], batch_size=1, max_gen_toks=512)
        if t is not None:
            musique_tasks.append(t)

    if musique_tasks:
        log(f"Waiting for {len(musique_tasks)} MuSiQue tasks...")
        wait_for_tasks(musique_tasks)
        log("MuSiQue tasks completed")

    # Step 3: Wait for BBH if still running
    log("Checking BBH status...")
    while True:
        bbh_done = all(find_results(m, "bbh_cot_zeroshot") for m in ["BASELINE", "GPT_REWRITE", "LARGE_MODEL"])
        if bbh_done:
            log("All BBH tasks completed!")
            break
        log("BBH still running, checking again in 5 min...")
        time.sleep(300)

    # Step 4: Final status
    log("")
    log("=" * 60)
    log("FINAL RESULTS SUMMARY")
    log("=" * 60)
    for task in ["bbh_cot_zeroshot", "longbench_hotpotqa", "longbench_musique"]:
        log(f"\n--- {task} ---")
        for mode_key in ["BASELINE", "GPT_REWRITE", "LARGE_MODEL"]:
            r = find_results(mode_key, task)
            if r:
                metric_key, score = get_score(r[0], task)
                score_disp = f"{score*100:.2f}%" if score and score < 1 else f"{score}" if score else "?"
                log(f"  {mode_key:15s}: {score_disp} ({metric_key})")
            else:
                log(f"  {mode_key:15s}: NOT FOUND")

    # Step 5: Update XLSX
    log("\nUpdating XLSX...")
    update_xlsx()

    log("\nWatchdog finished!")
