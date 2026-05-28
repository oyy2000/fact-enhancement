#!/usr/bin/env python3
"""Fill E0-E3 Controls sheet: steering positive-trace and best-sweep generation token/ρ stats.

- ρ = total tokenizer tokens / N_steps, steps split by blank line (\\n\\n), same as Sec. 3.2 / Fig.1 code.
- Positive traces: rewritten JSON entries with exact_match==1, field resp_after (steering target text).
- Best sweep: per experiment prefix (e0, e1, …), maximize GSM8K flexible-extract accuracy; stats on
  paired samples jsonl, deduped by doc_id (keeps last row per id if lm_eval was rerun).
"""
from __future__ import annotations

import argparse
import json
import statistics
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from transformers import AutoTokenizer

ROOT = Path("/common/users/sl2148/Public/yang_ouyang/projects/fact-enhancement")
DEFAULT_XLSX = ROOT / "documents" / "rebuttal_experiment_data.xlsx"
CONTROL_MODEL = ROOT / "control_experiments" / "Qwen_Qwen2.5-3B-Instruct"
SWEEP_DIR = CONTROL_MODEL / "sweep_eval"

REWRITTEN = {
    "E0": CONTROL_MODEL / "rewritten_e0_paraphrase.json",
    "E1": CONTROL_MODEL / "rewritten_e1_random_step_compress.json",
    "E2": CONTROL_MODEL / "rewritten_e2_dense_incorrect.json",
    "E3": CONTROL_MODEL / "rewritten_e3_rule_based.json",
    "E3.2": CONTROL_MODEL / "rewritten_e3_2_gpt54mini.json",
}

SWEEP_PREFIX = {"E0": "e0", "E1": "e1", "E2": "e2", "E3": "e3", "E3.2": "e3.2"}

# Reference baselines (11_steer_all_datasets_exp.py: GPT_REWRITE L6 λ=4, LARGE_MODEL L6 λ=0.45)
EXPS_ROOT = ROOT / "exps"
NO_VECTOR_GSM8K = EXPS_ROOT / "no_vector" / "gsm8k_cot_zeroshot_unified" / "Qwen2.5-3B-Instruct_no_vector"
GPT_REWRITE_3B = EXPS_ROOT / "gpt_rewrites_unified_new" / "Qwen_Qwen2.5-3B-Instruct"
LARGE_MODEL_3B = EXPS_ROOT / "large_model_rewrites_unified_new" / "Qwen_Qwen2.5-3B-Instruct"
DENSE_REWRITTEN_JSON = GPT_REWRITE_3B / "rewritten_old.json"
INFAMILY_PAIRED_JSON = LARGE_MODEL_3B / "Qwen_Qwen2.5-7B-Instruct_paired_responses.json"


def count_tokens_rho(text: str, tokenizer) -> tuple[int, float, float]:
    steps = [s.strip() for s in (text or "").split("\n\n") if s.strip()]
    if not steps:
        return 1, 0.0, 0.0
    n = len(steps)
    lens = [len(tokenizer.encode(s, add_special_tokens=False)) for s in steps]
    tot = float(sum(lens))
    return n, tot, tot / n


def pos_trace_stats(path: Path, tokenizer) -> tuple[int, float, float] | None:
    if not path.is_file():
        return None
    data = json.loads(path.read_text(encoding="utf-8"))
    pos = [ex for ex in data if float(ex.get("exact_match", 0)) >= 1.0 - 1e-9]
    toks, rhos = [], []
    for ex in pos:
        _, tt, r = count_tokens_rho(ex.get("resp_after") or "", tokenizer)
        toks.append(tt)
        rhos.append(r)
    if not toks:
        return 0, 0.0, 0.0
    return len(pos), round(statistics.mean(toks), 2), round(statistics.mean(rhos), 2)


def latest_results(jobdir: Path) -> Path | None:
    js = list(jobdir.glob("**/results_*.json"))
    return max(js, key=lambda p: p.stat().st_mtime) if js else None


def samples_for_result(results_path: Path) -> Path | None:
    ts = results_path.stem.replace("results_", "")
    parent = results_path.parent
    cand = list(parent.glob(f"samples_*_{ts}.jsonl"))
    if cand:
        return cand[0]
    alls = list(parent.glob("samples_*.jsonl"))
    return max(alls, key=lambda p: p.stat().st_mtime) if alls else None


def best_sweep_results(exp_prefix: str) -> tuple[Path | None, Path | None, float, str | None]:
    best_acc = -1.0
    best_dir: Path | None = None
    best_rp = None
    for d in sorted(SWEEP_DIR.iterdir()):
        if not d.is_dir() or not d.name.startswith(exp_prefix + "_L6_lam"):
            continue
        rp = latest_results(d)
        if rp is None:
            continue
        acc = json.loads(rp.read_text(encoding="utf-8"))["results"]["gsm8k_cot_zeroshot_unified"][
            "exact_match,flexible-extract"
        ]
        if acc > best_acc:
            best_acc, best_dir, best_rp = acc, d, rp
    sp = samples_for_result(best_rp) if best_rp else None
    return best_rp, sp, best_acc, best_dir.name if best_dir else None


def gen_stats_full_dedup(samples_path: Path, tokenizer) -> tuple[int, float, float, float] | None:
    """Dedup by doc_id: n_docs, mean n_steps, mean total tokens, mean ρ."""
    if samples_path is None or not samples_path.is_file():
        return None
    by_doc: dict = {}
    with samples_path.open(encoding="utf-8") as f:
        for line in f:
            rec = json.loads(line)
            did = rec.get("doc_id")
            rs = rec.get("resps") or []
            text = (rs[0][0] if rs and rs[0] else "") or ""
            n_s, tt, r = count_tokens_rho(text, tokenizer)
            by_doc[did] = (n_s, tt, r)
    if not by_doc:
        return 0, 0.0, 0.0, 0.0
    steps = [v[0] for v in by_doc.values()]
    toks = [v[1] for v in by_doc.values()]
    rhos = [v[2] for v in by_doc.values()]
    return (
        len(by_doc),
        round(statistics.mean(steps), 2),
        round(statistics.mean(toks), 2),
        round(statistics.mean(rhos), 2),
    )


def gen_stats_dedup(samples_path: Path, tokenizer) -> tuple[int, float, float] | None:
    if samples_path is None or not samples_path.is_file():
        return None
    by_doc: dict = {}
    with samples_path.open(encoding="utf-8") as f:
        for line in f:
            rec = json.loads(line)
            did = rec.get("doc_id")
            rs = rec.get("resps") or []
            text = (rs[0][0] if rs and rs[0] else "") or ""
            _, tt, r = count_tokens_rho(text, tokenizer)
            by_doc[did] = (tt, r)
    if not by_doc:
        return 0, 0.0, 0.0
    toks = [v[0] for v in by_doc.values()]
    rhos = [v[1] for v in by_doc.values()]
    return len(by_doc), round(statistics.mean(toks), 2), round(statistics.mean(rhos), 2)


def pick_gsm8k_samples_under(root: Path, subglob: str) -> Path | None:
    """First match under root for subglob, e.g. '**/Qwen2.5-3B-Instruct_L6_lam4p0/**/samples_gsm8k*.jsonl'."""
    matches = sorted(root.glob(subglob), key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0] if matches else None


def pick_no_vector_gsm8k() -> Path | None:
    if not NO_VECTOR_GSM8K.is_dir():
        return None
    js = sorted(NO_VECTOR_GSM8K.rglob("samples_gsm8k_cot_zeroshot_unified_*.jsonl"), key=lambda p: p.stat().st_mtime, reverse=True)
    return js[0] if js else None


def collect_reference_baselines(tokenizer) -> dict[str, dict]:
    """Keys: baseline, densesteer, infamily. Each has gen_full, pos (optional), layer, lam."""
    out: dict[str, dict] = {}

    p0 = pick_no_vector_gsm8k()
    g0 = gen_stats_full_dedup(p0, tokenizer) if p0 else None
    out["baseline"] = {
        "samples": p0,
        "gen_full": g0,
        "pos": None,
        "layer": None,
        "lam": None,
    }

    p1 = pick_gsm8k_samples_under(GPT_REWRITE_3B, "**/Qwen2.5-3B-Instruct_L6_lam4p0/**/samples_gsm8k*.jsonl")
    g1 = gen_stats_full_dedup(p1, tokenizer) if p1 else None
    ps1 = pos_trace_stats(DENSE_REWRITTEN_JSON, tokenizer) if DENSE_REWRITTEN_JSON.is_file() else None
    out["densesteer"] = {
        "samples": p1,
        "gen_full": g1,
        "pos": ps1,
        "layer": 6,
        "lam": 4.0,
    }

    p2 = pick_gsm8k_samples_under(LARGE_MODEL_3B, "**/Qwen2.5-3B-Instruct_L6_lam0p45/**/samples_gsm8k*.jsonl")
    g2 = gen_stats_full_dedup(p2, tokenizer) if p2 else None
    ps2 = pos_trace_stats(INFAMILY_PAIRED_JSON, tokenizer) if INFAMILY_PAIRED_JSON.is_file() else None
    out["infamily"] = {
        "samples": p2,
        "gen_full": g2,
        "pos": ps2,
        "layer": 6,
        "lam": 0.45,
    }
    return out


def collect_metrics(tokenizer) -> dict[str, tuple]:
    """exp_id -> (pos_n, pos_tok, pos_rho, gen_n, gen_tok, gen_rho, best_acc_pct, best_dir_name)."""
    out = {}
    for exp, prefix in SWEEP_PREFIX.items():
        rw = REWRITTEN.get(exp)
        ps = pos_trace_stats(rw, tokenizer) if rw else None
        _, sp, acc, bd = best_sweep_results(prefix)
        gs = gen_stats_dedup(sp, tokenizer) if sp else None
        pos_n, pos_tok, pos_rho = ps if ps else (None, None, None)
        if gs:
            gen_n, gen_tok, gen_rho = gs
        else:
            gen_n, gen_tok, gen_rho = None, None, None
        out[exp] = (
            pos_n,
            pos_tok,
            pos_rho,
            gen_n,
            gen_tok,
            gen_rho,
            round(acc * 100, 2) if acc >= 0 else None,
            bd,
        )
    return out


def fill_sheet(wb: openpyxl.Workbook, tokenizer) -> None:
    name = "E0-E3 Controls"
    if name not in wb.sheetnames:
        raise SystemExit(f"Sheet {name!r} not found; have {wb.sheetnames}")
    ws = wb[name]
    metrics = collect_metrics(tokenizer)

    # Locate Notes column (header row 1)
    notes_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and str(v).strip().lower() == "notes":
            notes_col = c
            break
    if notes_col is None:
        notes_col = ws.max_column + 1

    # Avoid duplicating headers if re-run
    marker = ws.cell(1, 14).value
    if marker and str(marker).startswith("Steering+"):
        start_col = 14
    else:
        ws.insert_cols(notes_col, 4)
        start_col = notes_col
        headers = [
            "Steering+ Avg\nTotal Tok",
            "Steering+\nρ (tok/step)",
            "Best-λ Gen\nAvg Total Tok",
            "Best-λ Gen\nρ (tok/step)",
        ]
        for i, h in enumerate(headers):
            cell = ws.cell(1, start_col + i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Map data rows by Exp label in column A (only Qwen2.5-3B control runs; same Exp id used for Llama placeholder rows)
    qwen_label = "Qwen2.5-3B-Instruct"
    for row in range(2, ws.max_row + 1):
        exp = ws.cell(row, 1).value
        if exp not in metrics:
            continue
        model = ws.cell(row, 3).value
        if model != qwen_label:
            for j in range(4):
                ws.cell(row, start_col + j).value = None
            continue
        _, pos_tok, pos_rho, _, gen_tok, gen_rho, _, _ = metrics[exp]
        ws.cell(row, start_col, value=pos_tok)
        ws.cell(row, start_col + 1, value=pos_rho)
        ws.cell(row, start_col + 2, value=gen_tok)
        ws.cell(row, start_col + 3, value=gen_rho)

    # Reference rows: Baseline / DenseSteer / InFamilySteer (Qwen2.5-3B GSM8K)
    ref = collect_reference_baselines(tokenizer)
    col_layer, col_lam = 4, 5
    col_steps, col_rho_legacy = 12, 13

    def write_ref_row(row: int, key: str) -> None:
        block = ref.get(key) or {}
        gf = block.get("gen_full")
        pos = block.get("pos")
        if gf:
            n_d, avg_s, avg_tok, avg_r = gf
            ws.cell(row, col_steps, value=avg_s)
            ws.cell(row, col_rho_legacy, value=avg_r)
            if pos:
                _, ptok, prho = pos
                ws.cell(row, start_col, value=ptok)
                ws.cell(row, start_col + 1, value=prho)
            else:
                ws.cell(row, start_col, value=None)
                ws.cell(row, start_col + 1, value=None)
            ws.cell(row, start_col + 2, value=avg_tok)
            ws.cell(row, start_col + 3, value=avg_r)
        lyr, lm = block.get("layer"), block.get("lam")
        if lyr is not None:
            ws.cell(row, col_layer, value=lyr)
        if lm is not None:
            ws.cell(row, col_lam, value=lm)

    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 3).value != qwen_label:
            continue
        method = ws.cell(row, 2).value
        if not method:
            continue
        m = str(method)
        if "Baseline (No Steering)" in m:
            write_ref_row(row, "baseline")
        elif "DenseSteer (Best Config)" in m:
            write_ref_row(row, "densesteer")
        elif "InFamilySteer (Best Config)" in m:
            write_ref_row(row, "infamily")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--tokenizer", default="Qwen/Qwen2.5-3B-Instruct")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    print(f"Loading tokenizer {args.tokenizer} …")
    tokenizer = AutoTokenizer.from_pretrained(args.tokenizer, trust_remote_code=True)
    metrics = collect_metrics(tokenizer)
    for k, v in sorted(metrics.items()):
        print(k, v)
    ref = collect_reference_baselines(tokenizer)
    for k, v in ref.items():
        print("REF", k, v)

    if args.dry_run:
        return

    wb = openpyxl.load_workbook(args.xlsx)
    fill_sheet(wb, tokenizer)
    wb.save(args.xlsx)
    print(f"Wrote {args.xlsx}")


if __name__ == "__main__":
    main()
